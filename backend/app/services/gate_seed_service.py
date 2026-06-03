"""Seed assessment_gate_definitions from the checklist Excel file."""
import uuid
from pathlib import Path
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.assessment import AssessmentGateDefinition
from app.models.common import CommonSubcharacteristic

DATA_DIR_DOCKER = Path("/data")
DATA_DIR_DEV = Path(__file__).parent.parent.parent.parent / "data"
DATA_DIR = DATA_DIR_DOCKER if DATA_DIR_DOCKER.exists() else DATA_DIR_DEV

EXCEL_FILENAME = "PA33_Product_Assessment_Lifecycle_Checklist_Combined_UL_Texus_v0_1.xlsx"
SHEET_NAME = "03_QG_Check_Matrix"

COLUMN_MAP = {
    "Quality Gate": "gate_id",
    "Gate Name": "gate_name",
    "Lifecycle Phase": "lifecycle_phase",
    "Expected Maturity": "expected_maturity",
    "Quality Characteristic": "characteristic",
    "Quality Sub-characteristic": "subcharacteristic",
    "What to Check": "what_to_check",
    "Pass Criteria": "pass_criteria",
    "Required Evidence": "required_evidence",
    "Review Method": "review_method",
    "Blocking Level": "blocking_level",
    "Responsible Role": "responsible_role",
}


async def seed_gate_definitions(session: AsyncSession) -> bool:
    """Seed gate definitions from Excel. Returns True if seeded."""
    count_result = await session.execute(
        select(func.count()).select_from(AssessmentGateDefinition)
    )
    if count_result.scalar_one() > 0:
        return False

    excel_path = DATA_DIR / EXCEL_FILENAME
    if not excel_path.exists():
        return False

    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb[SHEET_NAME]
    except Exception:
        return False

    # Read headers from row 1
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[cell.column] = str(cell.value).strip()

    # Build subcharacteristic name → id lookup
    sc_result = await session.execute(select(CommonSubcharacteristic))
    subchar_lookup = {
        sc.name.lower().strip(): sc.id
        for sc in sc_result.scalars().all()
    }

    order = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        row_data = {}
        for col_idx, value in enumerate(row, start=1):
            header = headers.get(col_idx)
            if header and header in COLUMN_MAP:
                field = COLUMN_MAP[header]
                row_data[field] = str(value).strip() if value is not None else ""

        gate_id = row_data.get("gate_id", "").strip()
        if not gate_id or gate_id.lower() in ("quality gate", ""):
            continue

        subchar_name = row_data.get("subcharacteristic", "").lower().strip()
        subchar_id = subchar_lookup.get(subchar_name)

        session.add(AssessmentGateDefinition(
            id=str(uuid.uuid4()),
            gate_id=gate_id,
            gate_name=row_data.get("gate_name", ""),
            lifecycle_phase=row_data.get("lifecycle_phase", ""),
            expected_maturity=row_data.get("expected_maturity", ""),
            characteristic=row_data.get("characteristic", ""),
            subcharacteristic=row_data.get("subcharacteristic", ""),
            subchar_id=subchar_id,
            what_to_check=row_data.get("what_to_check", ""),
            pass_criteria=row_data.get("pass_criteria", ""),
            required_evidence=row_data.get("required_evidence") or None,
            review_method=row_data.get("review_method") or None,
            blocking_level=row_data.get("blocking_level", "P2"),
            responsible_role=row_data.get("responsible_role") or None,
            display_order=order,
        ))
        order += 1

    await session.commit()
    return True
